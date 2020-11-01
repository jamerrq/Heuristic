import pandas as pd
import numpy as np
from openpyxl import Workbook


def get_data(i):

    # Read data from sheet i
    df = pd.read_excel('Datos.xlsx', header=None, sheet_name='I' + str(i))
    # Get n, m, p
    n, m, p = map(int, df.loc[0][:3])

    # Create constrainsts
    consts = np.zeros((m, n + 1))
    # Read them
    for i in range(m):
        consts[i] = df.loc[i + 1].to_numpy()

    # Create fobs
    fobs = np.zeros((p, n))
    # Read them
    for i in range(p):
        fobs[i] = df.loc[i + 1 + m].to_numpy()[:-1]

    # Return all shit
    return n, m, p, consts, fobs


def constructive(i, niter=100000):
    # Get the data (from the previous method)
    n, m, p, consts, fobs = get_data(i)
    # Create the values to return (solution)
    values = np.ones(n)
    # To keep checking feasibility
    constemp = np.zeros(m)

    for i in range(m):
        for j in range(n):
            constemp[i] += consts[i, j]

    while True and niter:
        rest = -1
        # Looking for a non fullfilled constraint
        for i in range(m):
            if constemp[i] > consts[i, -1]:
                rest = i
                break

        # If we didn't find it (all constrainst are accomplished), we are done
        if rest == -1:
            break

        # With the not accomplished constraint,
        # we now look for the biggest coefficient
        maxi, index = -1, -1
        for i in range(n):
            if values[i] and consts[rest, i] > maxi:
                maxi = consts[rest, i]
                index = i

        # If we found it, we just remove it from the fob
        if index != -1:
            values[index] = 0
            # Remove its coeffiecients from the constraints
            for i in range(m):
                constemp[i] -= consts[i, index]

        niter -= 1

    # Get the fobs values
    fobs_values = np.zeros(p)
    for i in range(p):
        for j in range(n):
            fobs_values[i] += fobs[i, j] * values[j]

    #print(constemp)
    # Return all shit
    return(sum([1 - (constemp[i] <= consts[i, -1]) for i in range(m)]), m)
    #return values, n, constemp, m, fobs_values, p


# Create an excel book
wb = Workbook()

def write_solution(i):
    global wb
    values, n, constemp, m, fobs_values, p = constructive(i)
    wsi = wb.create_sheet('I' + str(i))
    wsi['A1'] = 1
    wsi['A2'] = sum(values)
    current = 1
    for col in range(n):
        if values[col]:
            cell = wsi.cell(row=3, column=current)
            cell.value = col + 1
            current += 1
    for col in range(m):
        cell = wsi.cell(row=4, column=col + 1)
        cell.value = constemp[col]
    for col in range(p):
        cell = wsi.cell(row=5, column=col + 1)
        cell.value = fobs_values[col]


# for i in range(20):
#     write_solution(i + 1)


# sheet = wb['Sheet']
# wb.remove_sheet(sheet)
# wb.save('Resultados.xlsx')
# wb.close()

# constructive(1,niter=10000)
#for i in range(20):
#    constructive(i + 1)
