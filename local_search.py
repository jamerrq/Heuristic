import pandas as pd
import numpy as np
import random as rd
from openpyxl import Workbook
import math
import signal
import os
import time


#rd.seed(0)

def get_data(i):

    # Read data from sheet i
    df = pd.read_excel('Datos.xlsx', header=None, sheet_name='I' + str(i))

    # Get n, m, p, n items, m restrictions, p fobs
    n, m, p = map(int, df.loc[0][:3])

    # Create constrainsts
    consts = np.zeros((m, n + 1))
    # Read them
    for i in range(m):
        consts[i] = df.loc[i + 1].to_numpy()

    # Create fobs
    fobs = np.zeros((p, n))
    # Read them
    for i in range(p):
        fobs[i] = df.loc[i + 1 + m].to_numpy()[:-1]

    # Return all shit
    return n, m, p, consts, fobs


def greedyRan(case, alpha):

    # Get the data (from the previous method)
    n, m, p, consts, fobs = get_data(case)

    left = np.zeros((m, n))
    right = np.zeros(m)
    for i in range(m):
        left[i] = consts[i][:-1]
        right[i] = consts[i][-1]

    # Create the values to return (solution)
    items = np.zeros(n)

    # Create calification value/weight
    cals = [None]*n
    for l in range(n):
        v = 0
        w = 0
        for j in range(p):
            v += fobs[j][l]
        for k in range(m):
            if False:
                w += 1 / abs(consts[k][l])
            else:
                w += consts[k][l]

        cals[l] = [v/w, l, False] # [score/grade index taken]

    # Sort list descending
    cals_sorted = sorted(cals, reverse=True)
    # Generate initial RCL
    RCL = cals_sorted[0 : int(len(cals) * alpha)]

    # Fix negative constraints
    fix = np.zeros(m)
    fix2 = np.zeros(m)
    for i in range(m):
        if right[i] < 0:
            fix[i] = sum(left[i][j] for j in range(n))
            fix2[i] = right[i]


    valObj = np.dot(left, items)

    # Make solution
    while RCL:

        # Choose a random item from the RCL
        index = rd.randint(0, len(RCL) - 1)
        # Modify its value in the solution
        items[RCL[index][1]] = 1


        # Calculate constraints
        valObj = np.dot(left, items)


        # If we got a wrong one
        if sum([1 - (valObj[i] <= right[i] - fix[i]) for i in range(m)]):
            items[RCL[index][1]] = 0
            valObj = np.dot(left, items)
            break

        # When an item is chosen
        real_index = RCL[index][1]
        cals[real_index][2] = True


        # Update RCL
        nottaken = [cals[i] for i in range(len(cals)) if not cals[i][2]]
        cals_sorted = sorted(nottaken, reverse=True)
        RCL = cals_sorted[0:int(len(cals_sorted)*alpha)]

    #print(sum([1 - (valObj[i] <= right[i]) for i in range(m)]))

    for j in range(len(cals)):
        if not cals[j][2]:
            items[cals[j][1]] = 1
            valObj = np.dot(left, items)
            if sum([1 - (valObj[i] <= right[i]) for i in range(m)]):
                items[cals[j][1]] = 0
                valObj = np.dot(left, items)


    #print(valObj, right)
    #print(case, sum([1 - (valObj[i] <= right[i]) for i in range(m)]), m)
    return items, left, right, fobs


def neighbors1(s):
    # Separate indexes from zeros and ones
    zeros = [i for i in range(len(s)) if 1 - s[i]]
    ones  = [i for i in range(len(s)) if s[i]]
    # Create the neighborhood with the original solution
    neighbors = []
    #
    for i in range(len(zeros)):
        for j in range(len(ones)):
            neighbor = s.copy()
            neighbor[zeros[i]] = 1
            neighbor[ones[j]]  = 0
            neighbors.append(neighbor)

    # for neighbor in neighbors:
    #     print(neighbor)
    return neighbors


def neighbors2(s):
    neighbors = [s]
    #
    for i in range(len(s)):
        neighbor = s.copy()
        neighbor[i] = 1 - neighbor[i]
        neighbors.append(neighbor)
    #
    #for neighbor in neighbors:
    #    print(neighbor)
    return neighbors


def neighbors3(s):
    # Number of neighbors is the number of variables
    nn = len(s)
    # Neighbor start with the original solution
    neighborhood = [s]
    # Create a set to avoid repetitions
    check = set()
    # Add the original
    check.add(''.join([str(x) for x in s]))
    # Fill the neighborhood
    while len(neighborhood) <= nn:
        # Choose a random number of zeros and ones
        ones = rd.randint(1, len(s))
        zers = len(s) - ones
        # Create a vector of ones and zeros
        neighbor = [0] * zers + [1] * ones
        # Shuffle the array
        rd.shuffle(neighbor)
        # Create the key to check if it is repeated
        key = ''.join([str(x) for x in neighbor])
        if not key in check:
            neighborhood.append(neighbor)
            check.add(key)

    # Print that shit
    #for neighbor in neighborhood:
    #    print(neighbor)

    return neighborhood


def neighbor_value(neighbor, left, right, fobs):
    m = len(left)
    #
    valObj = np.dot(left, neighbor)
    noFill = sum([1 - (valObj[i] <= right[i]) for i in range(m)])
    #
    sumFob = sum(np.dot(fobs, neighbor))
    #
    return noFill, -sumFob


def vnd(i, alpha=0.3, file=None):
    now = time.time()
    #
    s, l, r, fobs = greedyRan(i, alpha)
    s_value = neighbor_value(s, l, r, fobs)
    #
    if file:
        file.write('First: ' + str(s_value) + '\n')
    else:
        print('First:', neighbor_value(s, l, r, fobs))
    #
    j = 0
    #
    neighborhoods = [neighbors3, neighbors2, neighbors1]
    nn = len(neighborhoods)
    #
    last_neighbors = []
    last_value = []

    neighbors = []
    best_neig = s
    a = True
    while j < nn and a:
        neighbors = neighborhoods[j](s)
        neighbors.sort(key=lambda x:neighbor_value(x, l, r, fobs))
        best_neig = neighbors[0]
        #
        best_neig_value = neighbor_value(best_neig, l, r, fobs)
        s_value = neighbor_value(s, l, r, fobs)
        if best_neig_value < s_value:
            if file:
                file.write('Got better! ' + str(s_value) + ' ' \
                    + str(best_neig_value) + '\n')
            else:
                print('Got better! ', s_value, best_neig_value)
            j = 0
            s = best_neig
            last_neighbors = neighbors
            last_value = best_neig_value
        else:
            j += 1

        then = time.time()
        if then - now > 270:
            a = False

    ss = [x for x in last_neighbors if neighbor_value(x,l,r,fobs) == last_value]

    if file:
        file.write('Final: ' + str(s_value) + '\n')
    else:
        print('Final:', neighbor_value(s, l, r, fobs))

    return ss, len(s), [np.dot(l,s) for s in ss], len(l), \
        [np.dot(fobs,s) for s in ss], len(fobs), last_value[0]


# Initial temp T0, final temp TF, rc cooling factor, L length
def simulated_annealing(i, T0, TF, rc, L):
    now = time.time()
    s, l, r, fobs = greedyRan(i,0.6)
    s_value = neighbor_value(s, l, r, fobs)
    s_nei = None
    neighbors = []
    last_neighbors = []

    # while stop criteria 5minutes
    a = True
    while a:
        t = T0
        while t > TF and a:
            m = 0
            while m < L and a:
                m += 1
                neighbors = neighbors1(s)+neighbors2(s)
                neighbors.sort(key=lambda x:neighbor_value(x, l, r, fobs))
                s_nei = neighbors[0]
                s_nei_value = neighbor_value(s_nei, l, r, fobs)
                d = (s_value[1] - s_nei_value[1])*-1

                if s_value[0] - s_nei_value[0] > 0 or d<0:
                    #print('Got better! ', s_value, s_nei_value)
                    s = s_nei
                    last_neighbors = neighbors
                else:
                    try:
                        prob = math.exp(-d/t)
                    except Exception:
                        prob = 1
                    if rd.random() < prob:
                        s = s_nei
                        last_neighbors = neighbors

                s_value = neighbor_value(s,l,r,fobs)

                then = time.time()
                if then-now > 40:
                    a = False

            t = t*rc

    ss = [x for x in last_neighbors if neighbor_value(x,l,r,fobs) == s_value]

    return ss, len(s), [np.dot(l,s) for s in ss], len(l), \
        [np.dot(fobs,s) for s in ss], len(fobs), s_value[0]


def check_solution(i=1):
    n, m, _, consts, _ = get_data(i)
    left = np.zeros((m, n))
    right = np.zeros(m)
    #
    solution = []
    file = open('hamilton_solution.in')
    for line in file:
        row = list(map(int, line.split()))
        solution.extend(row)
    file.close()
    solution = np.array(solution)
    #
    for i in range(m):
        left[i] = consts[i][:-1]
        right[i] = consts[i][-1]

    valObj = np.dot(left, solution)
    return sum([1 - (valObj[i] <= right[i]) for i in range(m)])


def signal_handler(signum, frame):
    raise Exception("Timed out")


def print_results(i):
    #signal.signal(signal.SIGALRM, signal_handler)
    #signal.alarm(1000)   # Sixty seconds
    file = None
    try:
        try:
            file = open(f'./Solutions/sol{i}.log', 'w')
        except FileNotFoundError:
            os.system('mkdir ./Solutions')
            file = open(f'./Solutions/sol{i}.log', 'w')
        vnd(i,file=file)
    except Exception as e:
        print(str(e) + f'for case {i}')
        file.close()
    print(f'Finished case {i}!')
    file.close()


# Create an excel book
wb = Workbook()

def write_solution_vnd(i):
    global wb
    values, n, constemp, m, fobs_values, p, rnc = vnd(i)
    wsi = wb.create_sheet('I' + str(i))
    wsi['A1'] = len(values)
    wsi['C1'] = f"RNC = {rnc}"
    for i in range(len(values)):
        cell = wsi.cell(row = 5*i+2, column = 1)
        cell.value = sum(values[i])
        current = 1
        for col in range(n):
            if values[i][col]:
                cell = wsi.cell(row=5*i+3, column=current)
                cell.value = col + 1
                current += 1
        for col in range(m):
            cell = wsi.cell(row=5*i+4, column=col + 1)
            cell.value = constemp[i][col]
        for col in range(p):
            cell = wsi.cell(row=5*i+5, column=col + 1)
            cell.value = fobs_values[i][col]


for i in range(3):
    write_solution_vnd(i + 1)


del wb['Sheet']
wb.save('ResultadosVND.xlsx')
wb.close()

# Create an excel book
wb = Workbook()

def write_solution_s_a(i):
    global wb
    values, n, constemp, m, fobs_values, p, rnc = simulated_annealing(i,20,1,0.5,10)
    wsi = wb.create_sheet('I' + str(i))
    wsi['A1'] = len(values)
    wsi['C1'] = f"RNC = {rnc}"
    for i in range(len(values)):
        cell = wsi.cell(row = 5*i+2, column = 1)
        cell.value = sum(values[i])
        current = 1
        for col in range(n):
            if values[i][col]:
                cell = wsi.cell(row=5*i+3, column=current)
                cell.value = col + 1
                current += 1
        for col in range(m):
            cell = wsi.cell(row=5*i+4, column=col + 1)
            cell.value = constemp[i][col]
        for col in range(p):
            cell = wsi.cell(row=5*i+5, column=col + 1)
            cell.value = fobs_values[i][col]


for i in range(3):
    write_solution_s_a(i + 1)


del wb['Sheet']
wb.save('ResultadosSA.xlsx')
wb.close()
